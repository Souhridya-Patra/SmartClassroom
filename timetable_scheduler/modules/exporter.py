# modules/exporter.py
import os
import io
import pandas as pd
from datetime import datetime
from reportlab.lib          import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles   import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units     import cm
from reportlab.platypus      import (SimpleDocTemplate, Table, TableStyle,
                                     Paragraph, Spacer, HRFlowable)
from reportlab.lib.enums     import TA_CENTER, TA_LEFT
from openpyxl                import Workbook
from openpyxl.styles         import (Font, PatternFill, Alignment,
                                     Border, Side, GradientFill)
from openpyxl.utils          import get_column_letter
from db.connection           import execute_query
from config                  import WORKING_DAYS

# ── Output directory ────────────────────────────────────────
EXPORT_DIR = "exports"
os.makedirs(EXPORT_DIR, exist_ok=True)


# ════════════════════════════════════════════════════════════
#  SHARED DATA HELPERS
# ════════════════════════════════════════════════════════════

def get_all_slots():
    return execute_query(
        "SELECT * FROM time_slots ORDER BY slot_number", fetch=True
    )

def get_section_timetable(section_id: int):
    return execute_query("""
        SELECT tt.day_of_week, tt.session_type,
               ts.slot_number, ts.label AS slot_label,
               ts.start_time, ts.end_time, ts.is_break,
               sb.name AS subject_name, sb.code AS subject_code,
               t.full_name AS teacher_name,
               r.name AS room_name, r.room_type
        FROM timetable tt
        JOIN time_slots ts ON tt.slot_id    = ts.id
        JOIN subjects   sb ON tt.subject_id = sb.id
        JOIN teachers    t ON tt.teacher_id = t.id
        JOIN rooms       r ON tt.room_id    = r.id
        WHERE tt.section_id = %s
        ORDER BY ts.slot_number
    """, (section_id,), fetch=True)

def get_teacher_timetable(teacher_id: int):
    return execute_query("""
        SELECT tt.day_of_week, tt.session_type,
               ts.slot_number, ts.label AS slot_label,
               ts.start_time, ts.end_time,
               sb.name AS subject_name, sb.code AS subject_code,
               sc.name AS section_name, sm.sem_number,
               d.name AS dept_name,
               r.name AS room_name
        FROM timetable tt
        JOIN time_slots  ts ON tt.slot_id    = ts.id
        JOIN subjects    sb ON tt.subject_id = sb.id
        JOIN sections    sc ON tt.section_id = sc.id
        JOIN semesters   sm ON sc.sem_id     = sm.id
        JOIN departments  d ON sm.dept_id    = d.id
        JOIN rooms        r ON tt.room_id    = r.id
        WHERE tt.teacher_id = %s
        ORDER BY ts.slot_number
    """, (teacher_id,), fetch=True)

def build_grid_data(rows: list, all_slots: list, view: str) -> dict:
    """
    Returns {(slot_number, day): cell_text}
    view: 'section' or 'teacher'
    """
    lookup = {}
    for r in rows:
        day      = r["day_of_week"]
        slot_num = r["slot_number"]
        is_lab   = r.get("session_type") == "lab"
        tag      = "[LAB]" if is_lab else "[TH]"

        if view == "section":
            cell = f"{tag} {r['subject_code']}\n{r['teacher_name'].split()[0]}\n{r['room_name']}"
        else:
            cell = (f"{tag} {r['subject_code']}\n"
                    f"Sec {r['section_name']} Sem{r['sem_number']}\n"
                    f"{r['room_name']}")

        if (slot_num, day) not in lookup:
            lookup[(slot_num, day)] = cell
    return lookup


# ════════════════════════════════════════════════════════════
#  PDF EXPORT
# ════════════════════════════════════════════════════════════

# Day header colors
DAY_COLORS_PDF = {
    "Monday":    colors.HexColor("#2E86AB"),
    "Tuesday":   colors.HexColor("#7B68EE"),
    "Wednesday": colors.HexColor("#1ABC9C"),
    "Thursday":  colors.HexColor("#E67E22"),
    "Friday":    colors.HexColor("#C0392B"),
}

def _build_pdf_table_data(rows, all_slots, view):
    """Builds the 2D list for ReportLab Table."""
    teaching_slots = [s for s in all_slots if not s["is_break"]]
    all_slots_ordered = all_slots   # includes breaks

    lookup = build_grid_data(rows, teaching_slots, view)

    # Header row
    header = ["Time Slot"] + WORKING_DAYS
    table_data = [header]

    for slot in all_slots_ordered:
        start = str(slot["start_time"])[:5] if slot["start_time"] else ""
        end   = str(slot["end_time"])[:5]   if slot["end_time"]   else ""
        slot_label = f"{slot['label']}\n{start}–{end}"

        if slot["is_break"]:
            row = [slot_label] + ["— BREAK —"] * len(WORKING_DAYS)
        else:
            row = [slot_label]
            for day in WORKING_DAYS:
                row.append(lookup.get((slot["slot_number"], day), "—"))
        table_data.append(row)

    return table_data

def export_section_pdf(section_id: int, section_meta: dict) -> bytes:
    """Generates a PDF timetable for a section. Returns bytes."""
    rows      = get_section_timetable(section_id)
    all_slots = get_all_slots()
    table_data = _build_pdf_table_data(rows, all_slots, "section")

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        leftMargin=1.5*cm, rightMargin=1.5*cm,
        topMargin=2*cm,    bottomMargin=1.5*cm
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "title", parent=styles["Title"],
        fontSize=16, textColor=colors.HexColor("#2C3E50"),
        spaceAfter=4, alignment=TA_CENTER
    )
    sub_style = ParagraphStyle(
        "sub", parent=styles["Normal"],
        fontSize=10, textColor=colors.HexColor("#7F8C8D"),
        spaceAfter=2, alignment=TA_CENTER
    )

    dept    = section_meta.get("dept_name", "")
    sem     = section_meta.get("sem_number", "")
    sec     = section_meta.get("section_name", "")
    gen_at  = datetime.now().strftime("%d %b %Y, %I:%M %p")

    elements = [
        Paragraph(f"🎓 Institute Timetable", title_style),
        Paragraph(f"{dept} &nbsp;|&nbsp; Semester {sem} &nbsp;|&nbsp; Section {sec}", sub_style),
        Paragraph(f"Generated: {gen_at}", sub_style),
        Spacer(1, 0.4*cm),
        HRFlowable(width="100%", thickness=1.5,
                   color=colors.HexColor("#2C3E50")),
        Spacer(1, 0.3*cm),
    ]

    # Column widths
    page_w   = landscape(A4)[0] - 3*cm
    slot_col = 3.2*cm
    day_col  = (page_w - slot_col) / len(WORKING_DAYS)
    col_widths = [slot_col] + [day_col] * len(WORKING_DAYS)

    # Build table
    t = Table(table_data, colWidths=col_widths, repeatRows=1)

    # Base style
    style_cmds = [
        # Header row
        ("BACKGROUND",  (0,0), (-1,0), colors.HexColor("#2C3E50")),
        ("TEXTCOLOR",   (0,0), (-1,0), colors.white),
        ("FONTNAME",    (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE",    (0,0), (-1,0), 9),
        ("ALIGN",       (0,0), (-1,0), "CENTER"),
        ("VALIGN",      (0,0), (-1,0), "MIDDLE"),
        ("ROWBACKGROUNDS", (0,1), (-1,-1),
         [colors.HexColor("#F8F9FA"), colors.white]),
        # Slot column
        ("BACKGROUND",  (0,1), (0,-1), colors.HexColor("#ECF0F1")),
        ("FONTNAME",    (0,1), (0,-1), "Helvetica-Bold"),
        ("FONTSIZE",    (0,1), (0,-1), 8),
        # All cells
        ("FONTSIZE",    (1,1), (-1,-1), 7.5),
        ("ALIGN",       (0,0), (-1,-1), "CENTER"),
        ("VALIGN",      (0,0), (-1,-1), "MIDDLE"),
        ("GRID",        (0,0), (-1,-1), 0.5, colors.HexColor("#BDC3C7")),
        ("ROWBACKGROUNDS", (0,1), (-1,-1),
         [colors.HexColor("#F8F9FA"), colors.white]),
        ("TOPPADDING",  (0,0), (-1,-1), 5),
        ("BOTTOMPADDING",(0,0),(-1,-1), 5),
    ]

    # Day header colors
    for col_i, day in enumerate(WORKING_DAYS, start=1):
        style_cmds.append(
            ("BACKGROUND", (col_i,0), (col_i,0), DAY_COLORS_PDF[day])
        )

    # Break row highlighting
    for row_i, slot in enumerate(all_slots, start=1):
        if slot["is_break"]:
            style_cmds += [
                ("BACKGROUND", (0, row_i), (-1, row_i),
                 colors.HexColor("#FFF9E6")),
                ("TEXTCOLOR",  (0, row_i), (-1, row_i),
                 colors.HexColor("#888888")),
                ("FONTNAME",   (0, row_i), (-1, row_i),
                 "Helvetica-Oblique"),
            ]

    # Lab cell highlighting
    for row_i, slot in enumerate(all_slots, start=1):
        if slot["is_break"]:
            continue
        for col_i, day in enumerate(WORKING_DAYS, start=1):
            cell_val = table_data[row_i][col_i]
            if "[LAB]" in str(cell_val):
                style_cmds.append(
                    ("BACKGROUND", (col_i, row_i), (col_i, row_i),
                     colors.HexColor("#D5F5E3"))
                )
            elif "[TH]" in str(cell_val):
                style_cmds.append(
                    ("BACKGROUND", (col_i, row_i), (col_i, row_i),
                     colors.HexColor("#EBF5FB"))
                )

    t.setStyle(TableStyle(style_cmds))
    elements.append(t)

    # Legend
    elements += [
        Spacer(1, 0.5*cm),
        Paragraph(
            "<b>Legend:</b> &nbsp; "
            "<font color='#1a4a6e'>[TH] Theory</font> &nbsp;|&nbsp; "
            "<font color='#145a32'>[LAB] Lab/Practical</font> &nbsp;|&nbsp; "
            "— Free Period",
            ParagraphStyle("legend", fontSize=8,
                           textColor=colors.HexColor("#555"),
                           alignment=TA_LEFT)
        )
    ]

    doc.build(elements)
    buf.seek(0)
    return buf.read()


def export_teacher_pdf(teacher_id: int, teacher_name: str) -> bytes:
    """Generates a PDF timetable for a teacher."""
    rows       = get_teacher_timetable(teacher_id)
    all_slots  = get_all_slots()
    table_data = _build_pdf_table_data(rows, all_slots, "teacher")

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        leftMargin=1.5*cm, rightMargin=1.5*cm,
        topMargin=2*cm,    bottomMargin=1.5*cm
    )

    styles   = getSampleStyleSheet()
    gen_at   = datetime.now().strftime("%d %b %Y, %I:%M %p")
    elements = [
        Paragraph("🎓 Faculty Timetable", ParagraphStyle(
            "t", parent=styles["Title"], fontSize=16,
            textColor=colors.HexColor("#2C3E50"), alignment=TA_CENTER
        )),
        Paragraph(teacher_name, ParagraphStyle(
            "s", parent=styles["Normal"], fontSize=12,
            textColor=colors.HexColor("#2980B9"), alignment=TA_CENTER
        )),
        Paragraph(f"Generated: {gen_at}", ParagraphStyle(
            "g", parent=styles["Normal"], fontSize=9,
            textColor=colors.HexColor("#7F8C8D"), alignment=TA_CENTER
        )),
        Spacer(1, 0.4*cm),
        HRFlowable(width="100%", thickness=1.5,
                   color=colors.HexColor("#2C3E50")),
        Spacer(1, 0.3*cm),
    ]

    page_w   = landscape(A4)[0] - 3*cm
    slot_col = 3.2*cm
    day_col  = (page_w - slot_col) / len(WORKING_DAYS)
    col_widths = [slot_col] + [day_col] * len(WORKING_DAYS)

    t = Table(table_data, colWidths=col_widths, repeatRows=1)

    style_cmds = [
        ("BACKGROUND",  (0,0), (-1,0), colors.HexColor("#2C3E50")),
        ("TEXTCOLOR",   (0,0), (-1,0), colors.white),
        ("FONTNAME",    (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE",    (0,0), (-1,0), 9),
        ("ALIGN",       (0,0), (-1,-1), "CENTER"),
        ("VALIGN",      (0,0), (-1,-1), "MIDDLE"),
        ("GRID",        (0,0), (-1,-1), 0.5, colors.HexColor("#BDC3C7")),
        ("ROWBACKGROUNDS", (0,1), (-1,-1),
         [colors.HexColor("#F8F9FA"), colors.white]),
        ("BACKGROUND",  (0,1), (0,-1), colors.HexColor("#ECF0F1")),
        ("FONTNAME",    (0,1), (0,-1), "Helvetica-Bold"),
        ("FONTSIZE",    (0,1), (0,-1), 8),
        ("FONTSIZE",    (1,1), (-1,-1), 7.5),
        ("TOPPADDING",  (0,0), (-1,-1), 5),
        ("BOTTOMPADDING",(0,0),(-1,-1), 5),
    ]

    for col_i, day in enumerate(WORKING_DAYS, start=1):
        style_cmds.append(
            ("BACKGROUND", (col_i,0), (col_i,0), DAY_COLORS_PDF[day])
        )

    for row_i, slot in enumerate(all_slots, start=1):
        if slot["is_break"]:
            style_cmds += [
                ("BACKGROUND", (0,row_i), (-1,row_i),
                 colors.HexColor("#FFF9E6")),
                ("TEXTCOLOR",  (0,row_i), (-1,row_i),
                 colors.HexColor("#888888")),
                ("FONTNAME",   (0,row_i), (-1,row_i),
                 "Helvetica-Oblique"),
            ]
        else:
            for col_i, day in enumerate(WORKING_DAYS, start=1):
                cell_val = table_data[row_i][col_i]
                if "[LAB]" in str(cell_val):
                    style_cmds.append(
                        ("BACKGROUND", (col_i,row_i), (col_i,row_i),
                         colors.HexColor("#D5F5E3"))
                    )
                elif "[TH]" in str(cell_val):
                    style_cmds.append(
                        ("BACKGROUND", (col_i,row_i), (col_i,row_i),
                         colors.HexColor("#EBF5FB"))
                    )

    t.setStyle(TableStyle(style_cmds))
    elements.append(t)
    doc.build(elements)
    buf.seek(0)
    return buf.read()


# ════════════════════════════════════════════════════════════
#  EXCEL EXPORT
# ════════════════════════════════════════════════════════════

# Colors
XL_DAY_FILLS = {
    "Monday":    PatternFill("solid", fgColor="2E86AB"),
    "Tuesday":   PatternFill("solid", fgColor="7B68EE"),
    "Wednesday": PatternFill("solid", fgColor="1ABC9C"),
    "Thursday":  PatternFill("solid", fgColor="E67E22"),
    "Friday":    PatternFill("solid", fgColor="C0392B"),
}
XL_THEORY_FILL = PatternFill("solid", fgColor="EBF5FB")
XL_LAB_FILL    = PatternFill("solid", fgColor="D5F5E3")
XL_BREAK_FILL  = PatternFill("solid", fgColor="FFF9E6")
XL_SLOT_FILL   = PatternFill("solid", fgColor="ECF0F1")
XL_HEADER_FILL = PatternFill("solid", fgColor="2C3E50")

XL_THIN = Border(
    left=Side(style="thin",   color="BDC3C7"),
    right=Side(style="thin",  color="BDC3C7"),
    top=Side(style="thin",    color="BDC3C7"),
    bottom=Side(style="thin", color="BDC3C7"),
)
XL_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
XL_LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)


def _apply_xl_header(ws, title: str, subtitle: str, num_cols: int):
    """Writes a styled title block at the top of the sheet."""
    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1,   end_column=num_cols)
    ws.merge_cells(start_row=2, start_column=1,
                   end_row=2,   end_column=num_cols)
    ws.merge_cells(start_row=3, start_column=1,
                   end_row=3,   end_column=num_cols)

    c1 = ws.cell(1, 1, title)
    c1.font      = Font(bold=True, size=14, color="FFFFFF")
    c1.fill      = PatternFill("solid", fgColor="2C3E50")
    c1.alignment = XL_CENTER

    c2 = ws.cell(2, 1, subtitle)
    c2.font      = Font(bold=True, size=11, color="2C3E50")
    c2.alignment = XL_CENTER

    c3 = ws.cell(3, 1,
                 f"Generated: {datetime.now().strftime('%d %b %Y, %I:%M %p')}")
    c3.font      = Font(italic=True, size=9, color="7F8C8D")
    c3.alignment = XL_CENTER


def _write_xl_timetable(ws, rows, all_slots, view, start_row=5):
    """Writes the timetable grid into the worksheet."""
    lookup = build_grid_data(rows, all_slots, view)

    # Column header row
    ws.cell(start_row, 1, "Time Slot").font = Font(bold=True, color="FFFFFF")
    ws.cell(start_row, 1).fill             = XL_HEADER_FILL
    ws.cell(start_row, 1).alignment        = XL_CENTER
    ws.cell(start_row, 1).border           = XL_THIN

    for col_i, day in enumerate(WORKING_DAYS, start=2):
        c = ws.cell(start_row, col_i, day)
        c.font      = Font(bold=True, color="FFFFFF", size=10)
        c.fill      = XL_DAY_FILLS[day]
        c.alignment = XL_CENTER
        c.border    = XL_THIN

    # Data rows
    for row_offset, slot in enumerate(all_slots, start=1):
        r = start_row + row_offset
        start = str(slot["start_time"])[:5] if slot["start_time"] else ""
        end   = str(slot["end_time"])[:5]   if slot["end_time"]   else ""
        label = f"{slot['label']}\n{start}–{end}"

        # Slot label cell
        slot_cell = ws.cell(r, 1, label)
        slot_cell.fill      = XL_BREAK_FILL if slot["is_break"] else XL_SLOT_FILL
        slot_cell.font      = Font(bold=True, size=8,
                                   italic=bool(slot["is_break"]))
        slot_cell.alignment = XL_LEFT
        slot_cell.border    = XL_THIN
        ws.row_dimensions[r].height = 48

        for col_i, day in enumerate(WORKING_DAYS, start=2):
            c = ws.cell(r, col_i)
            c.border    = XL_THIN
            c.alignment = XL_CENTER

            if slot["is_break"]:
                c.value = "BREAK"
                c.fill  = XL_BREAK_FILL
                c.font  = Font(italic=True, color="888888", size=8)
            else:
                val = lookup.get((slot["slot_number"], day), "—")
                c.value = val.replace("[TH] ","").replace("[LAB] ","")

                if "[LAB]" in val:
                    c.fill = XL_LAB_FILL
                    c.font = Font(size=8, color="145A32")
                elif "[TH]" in val:
                    c.fill = XL_THEORY_FILL
                    c.font = Font(size=8, color="1A4A6E")
                else:
                    c.font = Font(size=8, color="AAAAAA")

    # Column widths
    ws.column_dimensions["A"].width = 18
    for col_i in range(2, len(WORKING_DAYS) + 2):
        ws.column_dimensions[get_column_letter(col_i)].width = 22


def export_section_excel(section_id: int, section_meta: dict) -> bytes:
    """Generates Excel timetable for a section."""
    rows      = get_section_timetable(section_id)
    all_slots = get_all_slots()

    wb = Workbook()
    ws = wb.active
    ws.title = f"Sec {section_meta.get('section_name','')}"

    dept = section_meta.get("dept_name", "")
    sem  = section_meta.get("sem_number", "")
    sec  = section_meta.get("section_name", "")

    _apply_xl_header(
        ws,
        title    = "Institute Timetable",
        subtitle = f"{dept}  |  Semester {sem}  |  Section {sec}",
        num_cols = len(WORKING_DAYS) + 1
    )
    _write_xl_timetable(ws, rows, all_slots, "section")

    # Legend sheet
    ls = wb.create_sheet("Legend")
    legend_data = [
        ("Color",          "Meaning"),
        ("Blue cell",      "Theory lecture"),
        ("Green cell",     "Lab / Practical"),
        ("Yellow cell",    "Break period"),
        ("Grey slot col",  "Time slot label"),
    ]
    for i, (k, v) in enumerate(legend_data, start=1):
        ls.cell(i, 1, k).font = Font(bold=(i==1))
        ls.cell(i, 2, v).font = Font(bold=(i==1))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def export_teacher_excel(teacher_id: int, teacher_name: str) -> bytes:
    """Generates Excel timetable for a teacher."""
    rows      = get_teacher_timetable(teacher_id)
    all_slots = get_all_slots()

    wb = Workbook()
    ws = wb.active
    ws.title = teacher_name[:28]

    _apply_xl_header(
        ws,
        title    = "Faculty Timetable",
        subtitle = teacher_name,
        num_cols = len(WORKING_DAYS) + 1
    )
    _write_xl_timetable(ws, rows, all_slots, "teacher")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def export_all_sections_excel() -> bytes:
    """
    Generates a single Excel workbook with one sheet per section.
    Useful for printing the entire institute timetable at once.
    """
    sections = execute_query("""
        SELECT sc.id, sc.name AS section_name,
               sm.sem_number, d.name AS dept_name
        FROM sections sc
        JOIN semesters sm ON sc.sem_id  = sm.id
        JOIN departments d ON sm.dept_id = d.id
        ORDER BY d.name, sm.sem_number, sc.name
    """, fetch=True)

    all_slots = get_all_slots()
    wb = Workbook()
    wb.remove(wb.active)   # remove default blank sheet

    for sec in sections:
        rows = get_section_timetable(sec["id"])
        if not rows:
            continue

        sheet_name = (f"{sec['dept_name'][:8]}_S{sec['sem_number']}"
                      f"_{sec['section_name']}")[:31]
        ws = wb.create_sheet(title=sheet_name)

        _apply_xl_header(
            ws,
            title    = "Institute Timetable",
            subtitle = (f"{sec['dept_name']}  |  "
                        f"Semester {sec['sem_number']}  |  "
                        f"Section {sec['section_name']}"),
            num_cols = len(WORKING_DAYS) + 1
        )
        _write_xl_timetable(ws, rows, all_slots, "section")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def export_all_teachers_excel() -> bytes:
    """Single workbook with one sheet per teacher."""
    teachers  = execute_query(
        "SELECT id, full_name FROM teachers ORDER BY full_name", fetch=True
    )
    all_slots = get_all_slots()
    wb        = Workbook()
    wb.remove(wb.active)

    for t in teachers:
        rows = get_teacher_timetable(t["id"])
        if not rows:
            continue
        ws = wb.create_sheet(title=t["full_name"][:31])
        _apply_xl_header(
            ws,
            title    = "Faculty Timetable",
            subtitle = t["full_name"],
            num_cols = len(WORKING_DAYS) + 1
        )
        _write_xl_timetable(ws, rows, all_slots, "teacher")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()